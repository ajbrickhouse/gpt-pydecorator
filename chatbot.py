#!/usr/bin/env python3
from openai_decorator.openai_decorator import openaifunc, get_openai_funcs

import openai
import os
import sys
import json
from dotenv import load_dotenv
import openpyxl
from datetime import datetime
import discord
import asyncio
import warnings
import sqlite3
import pandas as pd
import datetime

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

database = "chatbot.db"

# Load the API key from .env file
load_dotenv()
def default(o):
    if isinstance(o, datetime):
        return o.__str__()
    return str(o)  # Ensure content is always a string

json.JSONEncoder.default = default

openai.api_key = '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'

@openaifunc
def search_master_list(keyword: str) -> dict:
    """
    Searches the master list for any keyword and returns all of the information related to the matches
    @param keyword: The keyword to search for
    """
    print("Searching system list for keyword: " + keyword)
    wb = openpyxl.load_workbook('ACvFLO master list.xlsx', read_only=True, data_only=True)

    result = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []

        for row in sheet.iter_rows(values_only=True):
            if keyword in row:
                rows.append(row)

        if rows:
            header = [cell.value for cell in sheet[1]]
            rows.insert(0, header)

        result[sheet_name] = rows

    return result

@openaifunc
def search_master_schedule(keyword: str) -> dict:
    """
    Searches the schedule for any keyword and returns all of the information related to the matches
    @param keyword: The keyword to search for
    """
    print("Searching master schedule for keyword: " + keyword)
    try:
        wb = openpyxl.load_workbook('Master Schedule.xlsx', read_only=True, data_only=True)
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Invalid file format. Please make sure the file is in the correct format.")
        return {}
    except Exception as e:
        print(f"Unexpected error occurred while loading the workbook: {e}")
        return {}

    result = {}

    sheet_name = 'Master Schedule'  # Specify the sheet name to search

    try:
        sheet = wb[sheet_name]
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        return {}

    rows = []

    for row in sheet.iter_rows(values_only=True):
        if keyword in row:
            rows.append(row)

    if rows:
        header = [cell.value for cell in sheet[1]]
        rows.insert(0, header)

    result[sheet_name] = rows

    return result

@openaifunc
def search_all_files(keyword: str) -> dict:
    """
    Searches all files for any keyword and returns all of the information related to the matches
    @param keyword: The keyword to search for
    """
    print("Searching all files for keyword: " + keyword)
    master_list_result = search_master_list(keyword)
    master_schedule_result = search_master_schedule(keyword)

    result = {**master_list_result, **master_schedule_result}

    return result

@openaifunc
def add_to_table(content: str, author: str, table: str = "messages") -> str:
    """
    Adds a message to the database
    @param content: The content of the message
    @param author: The author of the message
    @param table: The table to add the message to
    """
    print(f"Adding message to table '{table}' \n {author} \n {content}")
    conn = sqlite3.connect(database)
    c = conn.cursor()

    # Check if the table exists
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
    result = c.fetchone()

    if result is None:
        # Table does not exist, return the available tables
        c.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = c.fetchall()
        conn.close()
        return f"Table '{table}' does not exist. Available tables: {', '.join([t[0] for t in tables])}"

    # Insert the message into the database
    c.execute(f"INSERT INTO {table} (content, author, timestamp) VALUES (?, ?, ?)", (content, author, str(datetime.datetime.now())))

    conn.commit()
    conn.close()

    return "Message added to database."

@openaifunc
def remove_from_table(ids: list, table: str = "messages") -> str:
    """
    Removes messages from the database based on a list of IDs
    @param ids: The list of IDs of the messages to remove
    @param table: The table to remove the messages from
    """
    print(f"Removing messages with IDs {ids} from table {table}")
    conn = sqlite3.connect(database)
    c = conn.cursor()

    # Check if the table exists
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
    result = c.fetchone()

    if result is None:
        # Table does not exist, return the available tables
        c.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = c.fetchall()
        conn.close()
        return f"Table '{table}' does not exist. Available tables: {', '.join([t[0] for t in tables])}"

    # Delete the messages from the database
    for id in ids:
        c.execute(f"DELETE FROM {table} WHERE id=?", (id,))

    conn.commit()
    conn.close()

    return "Messages removed from database."

@openaifunc
def get_messages(table: str = "messages", limit: int = 10) -> dict:
    """
    Fetches a limited number of messages from the database and categorizes them into a dictionary.
    @param table: The table to fetch messages from.
    @param limit: The maximum number of messages to fetch.
    """
    print(f"Fetching messages from table {table}")
    # Connect to your database
    conn = sqlite3.connect(database)  # Replace with your database file or connection details
    cursor = conn.cursor()

    # Check if the table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
    result = cursor.fetchone()

    if result is None:
        # Table does not exist, return the available tables
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        conn.close()
        return f"Table '{table}' does not exist. Available tables: {', '.join([t[0] for t in tables])}"

    # Execute a query to fetch a limited number of messages in descending order of timestamp
    # Modify the query as per your database schema and requirements
    query = "SELECT id, content, author, timestamp FROM {} ORDER BY timestamp DESC LIMIT ?".format(table)  # Fix the query to include the table name
    cursor.execute(query, (limit,))  # Remove the 'table' parameter from the execute method
    fetched_messages = cursor.fetchall()

    conn.close()

    # Categorize messages
    categorized_messages = {}

    for id, content, author, timestamp in fetched_messages:
        if id not in categorized_messages:
            categorized_messages[id] = []

        categorized_messages[id].append(timestamp)
        categorized_messages[id].append(author)
        categorized_messages[id].append(content)

    if not categorized_messages:
        return "No messages found."
    
    # insert instructions into the categorized_messages dict
    instructions = "Provide the ID and the Content"
    categorized_messages["instructions"] = instructions

    return categorized_messages

# ------------------------------------

def create_database():
    """
    Creates the database if it doesn't already exist
    """
    print("Creating database...")
    conn = sqlite3.connect(database)
    c = conn.cursor()

    # Create the table if it doesn't already exist
    c.execute('''CREATE TABLE IF NOT EXISTS messages
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, content TEXT, author TEXT, timestamp TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS notes
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, content TEXT, author TEXT, timestamp TEXT)''')

    conn.commit()
    conn.close()

def convert_to_dict(messages):
    # Define the keys for the dictionary
    keys = ["id", "content", "author", "type", "timestamp"]

    # Convert each tuple to a dictionary
    messages_dict = [dict(zip(keys, message)) for message in messages]

    return messages_dict

def chunk_output(output: str, max_length: int = 1500) -> list:
    """
    Splits the output into chunks without breaking code encapsulations.
    @param output: The output to split
    @param max_length: The maximum length of each chunk
    """
    chunks = []
    parts = output.split("```")  # Splitting based on code block delimiter

    for part in parts:
        # append to chunks if part is less than max_length
        if len(part) < max_length:
            chunks.append(part)
        else:
            # split part into smaller chunks
            while len(part) > max_length:
                sub_part = part[:max_length]
                part = part[max_length:]
                chunks.append(sub_part)
            chunks.append(part)

    # create a new list to store valid chunks
    valid_chunks = []

    # go through each chunk, strip out the ` characters and see if it's empty. If it is, remove it from the list
    for chunk in chunks:
        # if the chunk is empty or only spaces, skip it
        if chunk.strip("`").strip(" ") == "":
            continue

        # make sure every list item starts and ends with ' ``` ', if it doesn't, add it
        if not chunk.startswith("```"):
            chunk = "" + chunk

        if not chunk.endswith("```"):
            chunk = chunk + ""

        valid_chunks.append(chunk)

    return valid_chunks

# ChatGPT API Function
def send_message(message, messages, model_arg="gpt-4-1106-preview"):
    messages.append(message)

    # Convert the 'content' of each message to a string if it is not already
    for msg in messages:
        if isinstance(msg.get('content'), dict):
            msg['content'] = json.dumps(msg['content'], default=str)

    try:
        response = openai.ChatCompletion.create(
            model=model_arg,
            messages=messages,
            functions=get_openai_funcs(),
            function_call="auto",
        )
    except openai.error.OpenAIError as e:
        print(f"OpenAIError: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected Error: {e}")
        sys.exit(1)

    messages.append(response["choices"][0]["message"])
    return messages

class MyClient(discord.Client):
    async def start_bot(self):
        create_database()
        while True:
            try:
                await self.start(os.getenv("DISCORD_TOKEN"))
            except discord.errors.ConnectionClosed as e:
                print(f"Connection closed: {e}, attempting to reconnect in 5 seconds.")
                await asyncio.sleep(5)  # wait before reconnecting
            except Exception as e:
                print(f"Unexpected error: {e}")
                break  # Exit the loop if the error is not related to connection

    async def on_ready(self):
        print(f'Logged on as {self.user}!')

    async def on_message(self, message):
        if message.author == self.user:
            return  # Avoid responding to the bot's own messages

        messages = []  # Initialize messages as an empty list

        # await message.channel.send("Working on it...")

        # fetch the messages from the channel
        async for msg in message.channel.history(limit=20):

            if msg.content:
                messages.append({"role": "user", "content": f"{msg.author}: {msg.content}"})

        # remove the last message from the list

        # reverse the list so that the messages are in chronological order
        messages.reverse()
        messages.pop()

        messages.append({
            "role": "system", 
            "content": """I am an assistant here to help you with a variety of work tasks, including programming, planning, and note-taking. I can also access information from the 'ACvFLO Master List' and 'Master Schedule.'
            Don't make assumptions about what values to plug into functions. Ask for clarification if a user request is ambiguous. 
            When you ask what I can do, I'll provide guidance using a list of my capabilities from get_openai_funcs().  """ + str(get_openai_funcs())
        })


        messages = send_message({
            "role": "user", 
            "content": message.content}, 
            messages
        )
        
        # If ChatGPT response is a function call, call the function
        if messages[-1].get("function_call"):
            function_name = messages[-1]["function_call"]["name"]
            arguments = json.loads(messages[-1]["function_call"]["arguments"])

            if function_name in globals() and callable(globals()[function_name]):
                # if function_name == "keyword_search_excel":
                #     keyword = arguments.get("keyword")
                #     if keyword:
                #         function_response = globals()[function_name](keyword)
                #     else:
                #         print("Keyword argument is missing for keyword_search_excel")
                #         await message.channel.send("Error: Keyword argument is missing.")
                #         return  # Exit the function early
                # else:
                function_response = globals()[function_name](**arguments)

                # Send function result back to ChatGPT
                messages = send_message(
                    {
                        "role": "function",
                        "name": function_name,
                        "content": function_response,
                    },
                    messages,
                    # "gpt-3.5-turbo-1106",
                )

        # Get the final response from ChatGPT
        final_response = messages[-1]["content"]

        responses = []

        responses = chunk_output(final_response)

        # Send final responses to Discord
        for response in responses:
            await message.channel.send(response)
        
intents = discord.Intents.default()
intents.message_content = True

client = MyClient(intents=intents)
asyncio.run(client.start_bot())