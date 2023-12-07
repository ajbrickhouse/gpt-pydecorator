import openpyxl

def keyword_search_excel(filename, keyword):
    wb = openpyxl.load_workbook(filename)
    result = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []

        for row in sheet.iter_rows(values_only=True):
            if keyword in row:
                rows.append(row)

        if rows:
            header = [cell.value for cell in sheet[1]]
            rows.insert(0, header)

        result[sheet_name] = rows

    return result

print(keyword_search_excel("ACvFLO master list.xlsx", "D22033-03A1"))